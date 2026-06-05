"""
Merge clean TSV files into one Excel file per year.

Reads from:  data/tsv_clean/{year}/{list_type}/page-XXXX.tsv
Writes to:   data/output/{year}.xlsx  (one sheet per list_type)

Usage:
    python -m rosenwald.postprocess.merge_to_excel
    python -m rosenwald.postprocess.merge_to_excel --year 1897
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

_PROJECT_ROOT = Path(__file__).resolve().parents[2]

# Output column schemas 
# (LLM columns + year + page at end)
SHEET_COLS: dict[str, list[str]] = {
    "paris_quartiers":  ["arrondissement", "quartier", "profession", "nom", "annee", "notes", "adresse", "horaires", "sexe", "year", "page"],
    "paris_rues":       ["rue", "nom", "annee", "notes", "adresse", "horaires", "sexe", "year", "page"],
    "deps_cantons":     ["departement", "arrondissement_dept", "canton", "profession", "nom", "annee", "notes", "adresse", "horaires", "sexe", "year", "page"],
    "seine_cantons":    ["departement", "arrondissement_dept", "canton", "profession", "nom", "annee", "notes", "adresse", "horaires", "sexe", "year", "page"],
    "specialists":      ["specialite", "nom", "annee", "notes", "adresse", "horaires", "sexe", "year", "page"],
    "thermal_spas":     ["station", "nom", "annee", "notes", "adresse", "horaires", "sexe", "year", "page"],
    "bienfaisance":     ["institution", "arrondissement", "nom", "year", "page"],
    "prefecture_seine": ["categorie", "nom", "year", "page"],
}

SHEET_ORDER = [
    "paris_quartiers", "paris_rues",
    "deps_cantons", "seine_cantons",
    "specialists", "bienfaisance", "prefecture_seine",
    "thermal_spas",
]


def _read_lt_dir(lt_dir: Path, n_cols: int) -> list[list[str]]:
    rows: list[list[str]] = []
    for tsv in sorted(lt_dir.glob("*.tsv")):
        try:
            for line in tsv.read_text(encoding="utf-8", errors="ignore").splitlines():
                if not line.strip():
                    continue
                parts = line.split("\t")
                parts = parts[:n_cols] + [""] * max(0, n_cols - len(parts))
                rows.append(parts)
        except Exception as e:
            print(f"    [WARN] {tsv.name}: {e}")
    return rows


def _write_sheet(ws, cols: list[str], rows: list[list[str]]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws.append(cols)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEEFF")
        cell.alignment = Alignment(wrap_text=False)

    for row in rows:
        ws.append(row)

    for col_idx in range(1, len(cols) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or ""))
             for r in range(1, min(len(rows) + 2, 502))),
            default=8,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def merge_year(year: int, clean_root: Path) -> None:
    year_dir = clean_root / str(year)
    if not year_dir.exists():
        print(f"  [SKIP] No clean TSVs for {year}")
        return

    out_path = _PROJECT_ROOT / "data" / "output" / f"{year}.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        import openpyxl

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        total_rows = 0

        for lt in SHEET_ORDER:
            lt_dir = year_dir / lt
            if not lt_dir.exists():
                continue
            cols = SHEET_COLS.get(lt)
            if not cols:
                continue
            rows = _read_lt_dir(lt_dir, len(cols))
            if not rows:
                continue
            ws = wb.create_sheet(title=lt[:31])
            _write_sheet(ws, cols, rows)
            total_rows += len(rows)
            print(f"    [{lt:<20}] {len(rows):>6} rows")

        if not wb.sheetnames:
            print(f"  [SKIP] No data found for {year}")
            return

        wb.save(out_path)
        print(f"  [OK] {out_path.name}  ({total_rows} rows, {len(wb.sheetnames)} sheets)")

    except ImportError:
        print("  [WARN] openpyxl not installed — pip install openpyxl")
        for lt in SHEET_ORDER:
            lt_dir = year_dir / lt
            if not lt_dir.exists():
                continue
            cols = SHEET_COLS.get(lt)
            if not cols:
                continue
            rows = _read_lt_dir(lt_dir, len(cols))
            if not rows:
                continue
            csv_path = out_path.parent / f"{year}_{lt}.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                csv.writer(f, delimiter=";").writerows([cols] + rows)
            print(f"    Saved: {csv_path.name}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--year", type=int, default=0)
    args = parser.parse_args()

    clean_root = _PROJECT_ROOT / "data" / "tsv_clean"
    if not clean_root.exists():
        print("[ERROR] data/tsv_clean/ not found — run restructure_tsv.py first.")
        sys.exit(1)

    years = sorted(clean_root.iterdir())
    if args.year:
        years = [y for y in years if y.name == str(args.year)]

    for year_dir in years:
        if not year_dir.is_dir():
            continue
        print(f"\n[YEAR] {year_dir.name}")
        merge_year(int(year_dir.name), clean_root)

    print("\nDone.")


if __name__ == "__main__":
    main()
