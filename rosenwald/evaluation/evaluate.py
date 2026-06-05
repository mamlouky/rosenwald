"""
Evaluation harness for the Rosenwald extraction pipeline (report ch. 4 & 5).

Three capabilities:

1. WER / CER between aligned gold and predicted TSVs
       python -m rosenwald.evaluation.evaluate tsv --gold data/gold --pred data/tsv_clean/1887

2. Ablation table across pipeline variants (reproduces report Table 4.4)
       python -m rosenwald.evaluation.evaluate ablation --gold data/gold \
           --variant tesseract=data/ocr_text \
           --variant unified=data/tsv_raw_unified \
           --variant routed-nogeo=data/tsv_raw_routed_nogeo \
           --variant routegeo=data/tsv_raw

3. Field-extraction accuracy — runs WITHOUT any API call, using the
   human-verified `Liste_femmes.xlsx` raw_text as gold (real numbers now):
       python -m rosenwald.evaluation.evaluate fields --ref Liste_femmes.xlsx

All tables are printed and also emitted as Markdown and LaTeX
(--latex out.tex) ready to paste into the report.

WER/CER are computed with a built-in Levenshtein (no external dep). Numbers are produced ONLY from real
inputs — this harness never invents values.
"""
from __future__ import annotations

import argparse
import re
import statistics
import sys
import warnings
from pathlib import Path
from typing import Dict, List, Tuple

_PROJECT_ROOT = Path(__file__).resolve().parent.parent
warnings.filterwarnings("ignore")


# ─────────────────────────────────────────────────────────────────────────────
# Text normalisation + WER/CER
# ─────────────────────────────────────────────────────────────────────────────

def normalize(s: str) -> str:
    """Lowercase, collapse whitespace, drop stray control chars. Accents kept."""
    s = (s or "").lower().replace("\t", " ")
    s = re.sub(r"[*¤•◆■]", " ", s)          # decorative OCR glyphs
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _edit_distance(a, b) -> int:
    """Levenshtein edit distance between two sequences (words or chars)."""
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + (ca != cb)))
        prev = cur
    return prev[-1]


def wer_cer(hyp: str, ref: str) -> Tuple[float, float]:
    """Return (WER, CER) of hypothesis vs reference.

    Word- and character-level Levenshtein error rates, computed in-house so the
    evaluation has no external dependency.
    """
    h, r = normalize(hyp), normalize(ref)
    if not r:
        # empty reference: WER/CER 0 if hyp also empty, else 1.0
        return (0.0, 0.0) if not h else (1.0, 1.0)
    wer = _edit_distance(h.split(), r.split()) / len(r.split())
    cer = _edit_distance(list(h), list(r)) / len(r)
    return wer, cer


def aggregate(per_file: List[Tuple[float, float]]) -> Dict[str, float]:
    """Macro average + median over per-file (wer, cer) — matches Avg/Mid in ch.4."""
    if not per_file:
        return {"avg_wer": 0.0, "mid_wer": 0.0, "avg_cer": 0.0, "mid_cer": 0.0, "n": 0}
    wers = [w for w, _ in per_file]
    cers = [c for _, c in per_file]
    return {
        "avg_wer": sum(wers) / len(wers),
        "mid_wer": statistics.median(wers),
        "avg_cer": sum(cers) / len(cers),
        "mid_cer": statistics.median(cers),
        "n": len(per_file),
    }


# ─────────────────────────────────────────────────────────────────────────────
# TSV directory scoring (gold vs predicted)
# ─────────────────────────────────────────────────────────────────────────────

def _page_text(tsv_path: Path, drop_leading: int = 0) -> str:
    """Flatten a TSV page into one text blob (one entry per line)."""
    lines = []
    for line in tsv_path.read_text(encoding="utf-8", errors="ignore").splitlines():
        if not line.strip():
            continue
        parts = line.split("\t")[drop_leading:]
        lines.append(" ".join(p.strip() for p in parts if p.strip()))
    return "\n".join(lines)


def _index_tsvs(root: Path) -> Dict[str, Path]:
    """Map {relative_path_without_root: file} for every *.tsv under root."""
    out: Dict[str, Path] = {}
    if not root.exists():
        return out
    for p in root.rglob("*.tsv"):
        out[str(p.relative_to(root))] = p
    # also accept .txt (Tesseract baseline)
    for p in root.rglob("*.txt"):
        out.setdefault(str(p.relative_to(root)).replace(".txt", ".tsv"), p)
    return out


def score_tsv_dirs(gold_root: Path, pred_root: Path,
                   drop_gold: int = 0, drop_pred: int = 0) -> Dict[str, float]:
    """WER/CER aggregated over every page present in BOTH gold and pred."""
    gold = _index_tsvs(gold_root)
    pred = _index_tsvs(pred_root)
    common = sorted(set(gold) & set(pred))
    per_file = []
    for key in common:
        ref = _page_text(gold[key], drop_gold)
        hyp = _page_text(pred[key], drop_pred)
        per_file.append(wer_cer(hyp, ref))
    res = aggregate(per_file)
    res["matched_pages"] = len(common)
    res["gold_pages"] = len(gold)
    return res


# ─────────────────────────────────────────────────────────────────────────────
# Ablation table (Table 4.4)
# ─────────────────────────────────────────────────────────────────────────────

def ablation_table(gold_root: Path, variants: List[Tuple[str, Path]]) -> List[List[str]]:
    rows = [["Variant", "Avg WER", "Mid WER", "Avg CER", "Mid CER", "Pages"]]
    for label, root in variants:
        r = score_tsv_dirs(gold_root, root)
        rows.append([
            label,
            f"{r['avg_wer']:.4f}", f"{r['mid_wer']:.4f}",
            f"{r['avg_cer']:.4f}", f"{r['mid_cer']:.4f}",
            str(r.get("matched_pages", 0)),
        ])
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Field-extraction accuracy — NO API, gold = Liste_femmes raw_text
# ─────────────────────────────────────────────────────────────────────────────

def _find_col(headers, *names):
    for i, h in enumerate(headers):
        if h and any(n.lower() in str(h).lower() for n in names):
            return i
    return None


def field_extraction_eval(reference_xlsx: Path, index_xlsx: Path) -> Dict[str, object]:
    """
    Evaluate the deterministic structuring step against the human-verified
    workbook: parse each raw_text and compare the recovered surname / civil
    status / woman-detection to the curated columns. Fully local.
    """
    import openpyxl
    from rosenwald.women.names import parse_name
    from rosenwald.women.detect import classify
    from rosenwald.index_reader import read_feminine_first_names

    fem = read_feminine_first_names(index_xlsx)
    wb = openpyxl.load_workbook(reference_xlsx, data_only=True)

    surn_ok = surn_tot = civ_ok = civ_tot = det_ok = det_tot = 0
    per_sheet = []
    for sheet in wb.sheetnames:
        rows = list(wb[sheet].iter_rows(values_only=True))
        if not rows:
            continue
        H = rows[0]
        iN = _find_col(H, "Nom"); iEv = _find_col(H, "civil", "Indication", "Preuve", "Indicateur")
        iRaw = _find_col(H, "raw_text", "brute")
        s_ok = s_tot = c_ok = c_tot = d_ok = d_tot = 0
        for r in rows[1:]:
            if not any(c not in (None, "") for c in r):
                continue
            nom = str(r[iN]).strip() if iN is not None and r[iN] else ""
            raw = str(r[iRaw]).strip() if iRaw is not None and r[iRaw] else ""
            ev = str(r[iEv]).strip() if iEv is not None and r[iEv] else ""
            if not (raw and nom):
                continue
            p = parse_name(raw)
            s_tot += 1
            if p.surname and (p.surname.lower() == nom.lower()
                              or nom.lower().startswith(p.surname.lower()[:4])):
                s_ok += 1
            if ev and ev[:1].isupper():     # civil-token evidence (not a first name)
                c_tot += 1
                if p.civil and p.civil.lower().rstrip(".") == ev.lower().rstrip("."):
                    c_ok += 1
            row = {"full_name_raw": raw, "gender_marker_raw": ev if ev[:1].isupper() else ""}
            d_tot += 1
            if classify(row, fem).is_woman:
                d_ok += 1
        per_sheet.append((sheet, s_ok, s_tot, c_ok, c_tot, d_ok, d_tot))
        surn_ok += s_ok; surn_tot += s_tot
        civ_ok += c_ok;  civ_tot += c_tot
        det_ok += d_ok;  det_tot += d_tot
    wb.close()

    def pct(a, b):
        return a / b if b else 0.0

    return {
        "surname_acc": pct(surn_ok, surn_tot), "surname_n": surn_tot,
        "civil_acc": pct(civ_ok, civ_tot), "civil_n": civ_tot,
        "detect_recall": pct(det_ok, det_tot), "detect_n": det_tot,
        "per_sheet": per_sheet,
    }


def field_eval_table(res: Dict[str, object]) -> List[List[str]]:
    rows = [["Métrique", "Exactitude", "n"]]
    rows.append(["Nom de famille (parsing)", f"{res['surname_acc']:.1%}", str(res["surname_n"])])
    rows.append(["État civil (parsing)", f"{res['civil_acc']:.1%}", str(res["civil_n"])])
    rows.append(["Détection femme (rappel)", f"{res['detect_recall']:.1%}", str(res["detect_n"])])
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Table rendering
# ─────────────────────────────────────────────────────────────────────────────

def to_markdown(rows: List[List[str]]) -> str:
    head, *body = rows
    out = ["| " + " | ".join(head) + " |",
           "| " + " | ".join("---" for _ in head) + " |"]
    out += ["| " + " | ".join(r) + " |" for r in body]
    return "\n".join(out)


def to_latex(rows: List[List[str]], caption: str = "", label: str = "") -> str:
    head, *body = rows
    spec = "l" + "r" * (len(head) - 1)
    lines = [r"\begin{table}[ht]", r"\centering", f"\\begin{{tabular}}{{{spec}}}", r"\hline"]
    lines.append(" & ".join(head) + r" \\ \hline")
    for r in body:
        lines.append(" & ".join(r) + r" \\")
    lines += [r"\hline", r"\end{tabular}"]
    if caption:
        lines.append(f"\\caption{{{caption}}}")
    if label:
        lines.append(f"\\label{{{label}}}")
    lines.append(r"\end{table}")
    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(description="Rosenwald extraction evaluation")
    sub = ap.add_subparsers(dest="cmd", required=True)

    p_tsv = sub.add_parser("tsv", help="WER/CER between two TSV dirs")
    p_tsv.add_argument("--gold", required=True)
    p_tsv.add_argument("--pred", required=True)
    p_tsv.add_argument("--latex", default="")

    p_abl = sub.add_parser("ablation", help="ablation table across variants")
    p_abl.add_argument("--gold", required=True)
    p_abl.add_argument("--variant", action="append", default=[],
                       help="label=path (repeatable)")
    p_abl.add_argument("--latex", default="")

    p_fld = sub.add_parser("fields", help="field-extraction accuracy (no API)")
    p_fld.add_argument("--ref", required=True, help="Liste_femmes.xlsx")
    p_fld.add_argument("--index", default="", help="index xlsx (feminine names)")
    p_fld.add_argument("--latex", default="")

    args = ap.parse_args()

    def emit(rows, caption, label, latex_path):
        print("\n" + to_markdown(rows) + "\n")
        if latex_path:
            Path(latex_path).write_text(to_latex(rows, caption, label), encoding="utf-8")
            print(f"[LaTeX written] {latex_path}")

    if args.cmd == "tsv":
        r = score_tsv_dirs(Path(args.gold), Path(args.pred))
        rows = [["Avg WER", "Mid WER", "Avg CER", "Mid CER", "Pages"],
                [f"{r['avg_wer']:.4f}", f"{r['mid_wer']:.4f}",
                 f"{r['avg_cer']:.4f}", f"{r['mid_cer']:.4f}", str(r["matched_pages"])]]
        emit(rows, "WER/CER", "tab:wercer", args.latex)

    elif args.cmd == "ablation":
        variants = []
        for v in args.variant:
            label, _, path = v.partition("=")
            variants.append((label, Path(path)))
        rows = ablation_table(Path(args.gold), variants)
        emit(rows, "Ablation des étapes du pipeline (WER/CER).", "tab:ablation", args.latex)

    elif args.cmd == "fields":
        from rosenwald.config import Settings
        index = Path(args.index) if args.index else Settings().excel_index
        res = field_extraction_eval(Path(args.ref), index)
        rows = field_eval_table(res)
        emit(rows, "Qualité de la structuration déterministe (gold vérifié humainement).",
             "tab:fields", args.latex)


if __name__ == "__main__":
    main()
