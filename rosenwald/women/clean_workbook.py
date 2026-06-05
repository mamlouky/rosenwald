"""
Clean the verified women workbook (reproducible, auditable).

    python -m rosenwald.women.clean_workbook            # -> data/reference/Liste_femmes_corrige.xlsx
    python -m rosenwald.women.clean_workbook --in X.xlsx --out Y.xlsx
"""
from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

from rosenwald.config import Settings

EVIDENCE_ALIASES = ("état civil", "etat civil", "indication", "preuve", "indicateur")
NEW_EVIDENCE = "Indicateur"
PLAUSIBLE = (1875, 1925)   
DOMINANCE = 0.60

# Entries verified by the historian to be men (unisex first name -> false positive).
# Keyed by (cluster_key(surname), normalised given).
CONFIRMED_FALSE_POSITIVES = {("leblanc", "camille")}


def _strip_accents(s: str) -> str:
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()


def cluster_key(surname: str) -> str:
    """Accent/spelling-insensitive key to group the same person across volumes."""
    s = _strip_accents(str(surname)).lower()
    s = re.sub(r"[^a-z ]", " ", s)
    s = re.sub(r"\b[il]\b", " ", s)          # stray single-letter "I"/"l" artefacts
    return re.sub(r"\s+", " ", s).strip()


def clean_surname(s: str) -> str:
    """Display-level surname cleanup (drop OCR junk, trailing title marks)."""
    s = str(s).replace("✓", "").replace("*", "").replace("☊", "")
    s = re.sub(r"\(\s*[Il]\s*$", "", s)       # trailing "( I"
    s = re.sub(r"\s+[Il]\s*$", "", s)          # trailing " I"
    return re.sub(r"\s+", " ", s).strip(" ,.(")


def to_year(v) -> int | None:
    m = re.search(r"\b(1[89]\d{2})\b", str(v))
    return int(m.group(1)) if m else None


def _find(headers, *aliases):
    for i, h in enumerate(headers):
        if h and any(a in str(h).lower() for a in aliases):
            return i
    return None


def clean(in_path: Path, out_path: Path, changelog_path: Path) -> dict:
    wb = openpyxl.load_workbook(in_path)
    changes = []

    # rename evidence header(s); drop duplicate evidence column 
    sheet_cols = {}     # sheet -> (i_nom, i_dip)
    for ws in wb.worksheets:
        headers = [c.value for c in ws[1]]
        evid_idxs = [i for i, h in enumerate(headers)
                     if h and any(a in str(h).lower() for a in EVIDENCE_ALIASES)]
        if evid_idxs:
            ws.cell(row=1, column=evid_idxs[0] + 1, value=NEW_EVIDENCE)
            for extra in sorted(evid_idxs[1:], reverse=True):   # drop duplicates
                ws.delete_cols(extra + 1)
        headers = [c.value for c in ws[1]]                       # refresh
        sheet_cols[ws.title] = (_find(headers, "nom"), _find(headers, "diplôme", "diplome"),
                                _find(headers, "prénom", "prenom"))

    # drop entries verified to be men (false positives)
    for ws in wb.worksheets:
        i_nom, i_dip, i_pre = sheet_cols[ws.title]
        if i_nom is None:
            continue
        to_del = []
        for r in range(2, ws.max_row + 1):
            sv = ws.cell(row=r, column=i_nom + 1).value
            if sv in (None, ""):
                continue
            gv = ws.cell(row=r, column=i_pre + 1).value if i_pre is not None else ""
            gkey = re.sub(r"[^a-z]", "", _strip_accents(str(gv or "")).lower())
            if (cluster_key(sv), gkey) in CONFIRMED_FALSE_POSITIVES:
                to_del.append((r, str(sv), str(gv)))
        for r, sv, gv in sorted(to_del, reverse=True):
            ws.delete_rows(r)
            changes.append((ws.title, r, "SUPPRIMÉE", f"{sv} ({gv})", "faux positif confirmé (homme)"))

    # build clusters across all sheets
    clusters = defaultdict(list)   # key -> list of (ws_title, rownum, surname, year)
    for ws in wb.worksheets:
        i_nom, i_dip, i_pre = sheet_cols[ws.title]
        if i_nom is None:
            continue
        for r in range(2, ws.max_row + 1):
            sv = ws.cell(row=r, column=i_nom + 1).value
            if sv in (None, ""):
                continue
            yr = to_year(ws.cell(row=r, column=i_dip + 1).value) if i_dip is not None else None
            clusters[cluster_key(sv)].append((ws.title, r, str(sv), yr))

    # resolve each cluster and apply 
    for key, members in clusters.items():
        if not key:
            continue
        # dominant clean surname spelling
        spell = Counter(clean_surname(s) for _, _, s, _ in members)
        modal_surname = spell.most_common(1)[0][0]
        plaus = [y for _, _, _, y in members if y and PLAUSIBLE[0] <= y <= PLAUSIBLE[1]]
        unify_year = None
        distinct = set(plaus)
        if len(distinct) == 1:
            unify_year = next(iter(distinct))         # one known year -> repair toward it
        elif plaus:
            yr, cnt = Counter(plaus).most_common(1)[0]
            if cnt >= max(2, DOMINANCE * len(plaus)):  # else ambiguous (possible homonyms)
                unify_year = yr

        for title, r, surname, yr in members:
            ws = wb[title]
            i_nom, i_dip, i_pre = sheet_cols[title]
            cleaned = clean_surname(surname)
            target = modal_surname if len(members) > 1 else cleaned
            if target and target != surname:
                ws.cell(row=r, column=i_nom + 1, value=target)
                changes.append((title, r, "Nom", surname, target))
            if unify_year is not None and i_dip is not None:
                cur = ws.cell(row=r, column=i_dip + 1).value
                cy = to_year(cur)
                cur_plausible = cy is not None and PLAUSIBLE[0] <= cy <= PLAUSIBLE[1]
                if cur not in (None, "") and not cur_plausible:
                    # only repair garbage years (a street number / hour / impossible
                    # value leaked in); never override an already-plausible year,
                    # which could belong to a distinct homonym.
                    ws.cell(row=r, column=i_dip + 1, value=unify_year)
                    changes.append((title, r, "Date de diplôme", cur, unify_year))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    with changelog_path.open("w", newline="", encoding="utf-8-sig") as fh:
        wr = csv.writer(fh)
        wr.writerow(["feuille", "ligne", "champ", "ancien", "nouveau"])
        wr.writerows(changes)

    by_field = Counter(c[2] for c in changes)
    return {"out": out_path, "changelog": changelog_path, "n_changes": len(changes),
            "by_field": dict(by_field), "n_clusters": sum(1 for k in clusters if k)}


def main() -> None:
    s = Settings()
    ap = argparse.ArgumentParser()
    # Always start from the pristine original so the changelog is complete.
    ap.add_argument("--in", dest="inp", default=str(s.data_dir / "reference" / "Liste_femmes.xlsx"))
    ap.add_argument("--out", default=str(s.data_dir / "reference" / "Liste_femmes_corrige.xlsx"))
    ap.add_argument("--changelog", default=str(s.data_dir / "reference" / "corrections_changelog.csv"))
    a = ap.parse_args()
    res = clean(Path(a.inp), Path(a.out), Path(a.changelog))
    print(f"[OK] {res['out']}")
    print(f"     clusters (personnes regroupées) : {res['n_clusters']}")
    print(f"     cellules corrigées             : {res['n_changes']}  {res['by_field']}")
    print(f"     journal                         : {res['changelog']}")


if __name__ == "__main__":
    main()
