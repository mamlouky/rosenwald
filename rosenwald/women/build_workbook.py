"""
Build the women-doctors workbook from the pipeline's merged output.

This is the final women-extraction deliverable. It consumes the merged,
woman-annotated TSVs produced by `women_filter.py`
(data/tsv_merged/{year}_all.tsv) and writes a single multi-sheet Excel
workbook — one sheet per geographic list type, with the schema used for
historical analysis:

    Nom(s) | Prénom(s) | [État civil] | Date de diplôme | <geo context> |
    [Adresse] | [Horaires] | Année du volume | Numéro de page |
    [Indication/Preuve] | Entrée brute (raw_text)

Each woman row carries the *evidence* token (the civil marker or feminine
first name that triggered detection) so every entry is auditable back to a
concrete cue in the scanned source. The workbook is the automatic baseline;
the historian then verifies and corrects the edge cases.

Usage:
    python -m rosenwald.women.build_workbook                       # all years -> data/output/Liste_femmes.xlsx
    python -m rosenwald.women.build_workbook --out Liste_femmes.xlsx
    python -m rosenwald.women.build_workbook --validate Liste_femmes.xlsx   # compare to a reference
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

_PROJECT_ROOT = Path(__file__).resolve().parent.parent

from rosenwald.config import Settings
from rosenwald.index_reader import read_feminine_first_names
from rosenwald.women.names import parse_name
from rosenwald.women.detect import classify

# Columns of the merged TSV 
MERGED_COLUMNS = [
    "year", "pdf_page", "list_type",
    "arrondissement", "quartier", "departement", "canton",
    "specialite", "station", "rue",
    "profession_section",
    "full_name_raw", "diploma_year",
    "address_raw", "phone_raw", "hours_raw", "specialties_raw",
    "gender_marker_raw", "maiden_name_raw",
    "notes_raw", "entry_raw",
    "is_woman",
]

# Per-sheet schema.
# Each entry is (column_header, source). `source` is either a merged-TSV key,
# or a special token resolved per row:
#   _surname  -> family name parsed from full_name_raw
#   _given    -> first name / initials parsed from full_name_raw
#   _evidence -> detection token (civil marker or feminine first name)
# list_type -> (sheet_name, columns). Order of this dict = sheet order.
SHEETS: Dict[str, Tuple[str, List[Tuple[str, str]]]] = {
    "paris_quartiers": ("paris_quartiers", [
        ("Nom(s)", "_surname"), ("Prénom(s)", "_given"), ("Indicateur", "_evidence"),
        ("Date de diplôme", "diploma_year"), ("Profession", "profession_section"),
        ("Arrondissement", "arrondissement"), ("Quartier", "quartier"),
        ("Adresse", "address_raw"), ("Horaires", "hours_raw"),
        ("Année du volume", "year"), ("Numéro de page", "pdf_page"),
        ("Entrée brute (raw_text)", "entry_raw"),
    ]),
    "deps_cantons": ("deps_cantons", [
        ("Nom(s)", "_surname"), ("Prénom(s)", "_given"),
        ("Date de diplôme", "diploma_year"), ("Profession", "profession_section"),
        ("Département", "departement"), ("Arrondissement", "arrondissement"),
        ("Canton", "canton"), ("Horaires", "hours_raw"),
        ("Année du volume", "year"), ("Numéro de page", "pdf_page"),
        ("Indicateur", "_evidence"), ("Entrée brute (raw_text)", "entry_raw"),
    ]),
    "paris_rues": ("paris_rues", [
        ("Nom(s)", "_surname"), ("Prénom(s)", "_given"), ("Indicateur", "_evidence"),
        ("Date de diplôme", "diploma_year"), ("Rue", "rue"),
        ("Année du volume", "year"), ("Numéro de page", "pdf_page"),
        ("Entrée brute (raw_text)", "entry_raw"),
    ]),
    "seine_cantons": ("seine_cantons", [
        ("Nom(s)", "_surname"), ("Prénom(s)", "_given"), ("Indicateur", "_evidence"),
        ("Date de diplôme", "diploma_year"), ("Profession", "profession_section"),
        ("Département", "departement"), ("Arrondissement", "arrondissement"),
        ("Canton", "canton"), ("Adresse", "address_raw"), ("Horaires", "hours_raw"),
        ("Année du volume", "year"), ("Numéro de page", "pdf_page"),
        ("Entrée brute (raw_text)", "entry_raw"),
    ]),
    "specialists": ("specialistes", [
        ("Nom(s)", "_surname"), ("Prénom(s)", "_given"),
        ("Date de diplôme", "diploma_year"), ("Spécialité", "specialite"),
        ("Année du volume", "year"), ("Numéro de page", "pdf_page"),
        ("Indicateur", "_evidence"), ("Entrée brute (raw_text)", "entry_raw"),
    ]),
    "thermal_spas": ("stations_thermales", [
        ("Nom(s)", "_surname"), ("Prénom(s)", "_given"),
        ("Date de diplôme", "diploma_year"), ("Station", "station"),
        ("Année du volume", "year"), ("Numéro de page", "pdf_page"),
        ("Indicateur", "_evidence"), ("Entrée brute (raw_text)", "entry_raw"),
    ]),
    "bienfaisance": ("bienfaisance", [
        ("Noms ", "_surname"), ("Arrondissement", "arrondissement"),
        ("Année", "year"), ("Page", "pdf_page"),
        ("Entrée brute (raw_text)", "entry_raw"),
    ]),
}


# Reading merged TSVs
def _iter_merged_rows(merged_dir: Path) -> Iterable[Dict[str, str]]:
    """Yield row dicts from every data/tsv_merged/*_all.tsv file."""
    files = sorted(merged_dir.glob("*_all.tsv"))
    if not files:
        raise FileNotFoundError(
            f"No merged TSVs found in {merged_dir}.\n"
            f"Run women_filter.py for each year first."
        )
    for f in files:
        with f.open(encoding="utf-8", newline="") as fh:
            reader = csv.DictReader(fh, delimiter="\t")
            for row in reader:
                yield {k: (v or "").strip() for k, v in row.items()}


def _resolve(row: Dict[str, str], source: str, parsed, evidence: str) -> str:
    if source == "_surname":
        return parsed.surname
    if source == "_given":
        return parsed.given
    if source == "_evidence":
        return evidence
    return row.get(source, "")


def build_rows(
    merged_rows: Iterable[Dict[str, str]],
    fem_names,
) -> Dict[str, List[List[str]]]:
    """Return {list_type: [row, ...]} of women rows in each sheet's schema."""
    sheet_rows: Dict[str, List[List[str]]] = {lt: [] for lt in SHEETS}
    for row in merged_rows:
        lt = row.get("list_type", "")
        if lt not in SHEETS:
            continue
        ev = classify(row, fem_names)
        if not ev.is_woman:
            continue
        parsed = parse_name(row.get("full_name_raw", ""))
        _, cols = SHEETS[lt]
        sheet_rows[lt].append(
            [_resolve(row, src, parsed, ev.evidence) for _, src in cols]
        )
    return sheet_rows


# Writing the workbook
def write_workbook(sheet_rows: Dict[str, List[List[str]]], out_path: Path) -> int:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    total = 0
    for lt, (sheet_name, cols) in SHEETS.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.append([h for h, _ in cols])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDEEFF")
        for r in sheet_rows.get(lt, []):
            ws.append(r)
        total += len(sheet_rows.get(lt, []))
        # best-effort column widths
        for ci in range(1, len(cols) + 1):
            letter = ws.cell(row=1, column=ci).column_letter
            ws.column_dimensions[letter].width = 18

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return total


def build(settings: Settings, out_path: Path) -> Path:
    fem_names = read_feminine_first_names(settings.excel_index)
    merged_dir = settings.data_dir / "tsv_merged"
    sheet_rows = build_rows(_iter_merged_rows(merged_dir), fem_names)

    total = write_workbook(sheet_rows, out_path)
    print(f"\n[OK] Women workbook written: {out_path}")
    for lt, (sheet_name, _) in SHEETS.items():
        print(f"     {sheet_name:20s} {len(sheet_rows.get(lt, [])):5d} women")
    print(f"     {'TOTAL':20s} {total:5d} women")
    return out_path


# Validation against a reference workbook (reproduction fidelity)
def validate(reference_path: Path, settings: Settings) -> None:
    """Compare the generated workbook against a reference, per sheet.

    Reports row-count agreement and surname-set overlap """
    import openpyxl

    fem_names = read_feminine_first_names(settings.excel_index)
    sheet_rows = build_rows(_iter_merged_rows(settings.data_dir / "tsv_merged"), fem_names)

    ref = openpyxl.load_workbook(reference_path, read_only=True, data_only=True)
    print(f"\n{'sheet':20s} {'gen':>6} {'ref':>6} {'shared surnames':>16}")
    print("-" * 52)
    for lt, (sheet_name, cols) in SHEETS.items():
        gen = sheet_rows.get(lt, [])
        gen_surn = {r[0].strip().lower() for r in gen if r and r[0]}
        ref_surn = set()
        if sheet_name in ref.sheetnames:
            ws = ref[sheet_name]
            for r in ws.iter_rows(min_row=2, values_only=True):
                if r and r[0]:
                    ref_surn.add(str(r[0]).strip().lower())
        inter = gen_surn & ref_surn
        union = gen_surn | ref_surn
        jac = len(inter) / len(union) if union else 0.0
        print(f"{sheet_name:20s} {len(gen):6d} {len(ref_surn):6d} {jac:15.1%}")
    ref.close()


# CLI
def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--out", default="", help="output xlsx path (default data/output/Liste_femmes.xlsx)")
    p.add_argument("--validate", default="", help="reference xlsx to compare against")
    args = p.parse_args()

    settings = Settings()
    if args.validate:
        validate(Path(args.validate), settings)
        return

    out = Path(args.out) if args.out else settings.data_dir / "output" / "Liste_femmes.xlsx"
    build(settings, out)


if __name__ == "__main__":
    main()
